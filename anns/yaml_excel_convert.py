import argparse
from pathlib import Path
import yaml
from openpyxl import Workbook, load_workbook

parser = argparse.ArgumentParser()
parser.add_argument("--src", type=Path, required=True)
parser.add_argument("--dst", type=Path, required=True)

def yaml_to_excel(yaml_fname, excel_fname):
    with open(yaml_fname) as f:
        objs = yaml.full_load(f)

    wb = Workbook()
    ws = wb.active

    for col, (cat, cat_objs) in enumerate(objs.items()):
        ws.cell(row=1, column=col+1).value = cat
        for row, obj in enumerate(cat_objs):
            ws.cell(row=row+2, column=col+1).value = obj

    wb.save(excel_fname)

def excel_to_yaml(excel_fname, yaml_fname):
    wb = load_workbook(excel_fname)
    ws = wb.active

    objs = {}
    for col in ws.columns:
        cat = col[0].value
        if cat is None or cat == "":
            break
        objs[cat] = []
        for elt in col[1:]:
            if elt.value is None or elt.value == "":
                break
            objs[cat].append(elt.value)

    with open(yaml_fname, "w") as f:
        yaml.dump(objs, f)

def main(args):
    if args.src.suffix == ".yaml":
        yaml_to_excel(args.src, args.dst)
    elif args.src.suffix == ".xlsx":
        excel_to_yaml(args.src, args.dst)
    else:
        raise ValueError

if __name__ == "__main__":
    args = parser.parse_args()
    main(args)
